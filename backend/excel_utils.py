import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def parse_excel_meta(file_path):
    """
    Extracts sheet names and unique columns across all sheets.
    """
    try:
        xl = pd.ExcelFile(file_path)
        sheet_names = xl.sheet_names
        
        all_columns = []
        for sheet in sheet_names:
            # Load only the header to extract columns quickly
            df = pd.read_excel(file_path, sheet_name=sheet, nrows=0)
            for col in df.columns:
                col_str = str(col).strip()
                if col_str not in all_columns:
                    all_columns.append(col_str)
                    
        return {
            "sheets": sheet_names,
            "columns": all_columns
        }
    except Exception as e:
        raise ValueError(f"Failed to read Excel file: {str(e)}")

def search_customer(file_path, search_term):
    """
    Searches for a term across all columns in all sheets.
    Returns combined DataFrame and a list of matched sheets.
    """
    try:
        all_sheets = pd.read_excel(file_path, sheet_name=None)
    except Exception as e:
        raise ValueError(f"Error reading Excel sheets: {str(e)}")

    if not all_sheets:
        raise ValueError("Excel file has no valid sheets.")

    results = []
    sheets_matched = []
    search_term_str = str(search_term).strip().lower()

    for sheet_name, df in all_sheets.items():
        if df.empty:
            continue
            
        # Create a boolean mask identifying rows where any cell contains the search term
        # Convert all cells to strings for uniform case-insensitive partial checking
        mask = df.astype(str).apply(
            lambda col: col.str.contains(search_term_str, case=False, na=False)
        ).any(axis=1)
        
        matches = df[mask]
        
        if not matches.empty:
            matches = matches.copy()
            # Insert Source Sheet column at the front
            matches.insert(0, "Source Sheet", sheet_name)
            results.append(matches)
            sheets_matched.append(sheet_name)
                
    if results:
        combined_df = pd.concat(results, ignore_index=True)
        # Replace NaN/NaT values with None on the combined DataFrame to make it JSON serializable
        combined_df = combined_df.where(pd.notnull(combined_df), None)
        return combined_df, sheets_matched
    
    return None, []

def save_and_style_excel(df, output_path):
    """
    Saves a dataframe to output_path and styles the Excel worksheet:
    - Bold header row with light blue background.
    - 'Source Sheet' column highlighted in light yellow.
    - Auto-fit column widths.
    """
    # Write to Excel
    df.to_excel(output_path, index=False)
    
    # Load with openpyxl to apply styling
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Font and styling configs
    header_font = Font(name='Segoe UI', size=11, bold=True, color='000000')
    header_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid') # Soft blue accent
    header_align = Alignment(horizontal='center', vertical='center')
    
    source_sheet_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') # Soft yellow
    regular_font = Font(name='Segoe UI', size=11)
    
    # Apply header styling
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        
    # Find which index the "Source Sheet" column has
    source_sheet_col_idx = 1
    for col_num in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col_num).value == "Source Sheet":
            source_sheet_col_idx = col_num
            break
            
    # Apply regular formatting and source sheet highlight
    for row_num in range(2, ws.max_row + 1):
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = regular_font
            if col_num == source_sheet_col_idx:
                cell.fill = source_sheet_fill
                
    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(output_path)
