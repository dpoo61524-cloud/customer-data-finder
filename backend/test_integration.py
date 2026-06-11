import requests
import os
import pandas as pd
from openpyxl import load_workbook

API_URL = "http://127.0.0.1:8000"
FILE_PATH = "../sample_customers.xlsx"

def run_tests():
    print("Starting Integration Tests...")
    
    # Check if sample file exists
    if not os.path.exists(FILE_PATH):
        print(f"Error: {FILE_PATH} not found. Generate mock data first.")
        return
        
    # 1. Test /upload
    print("\n1. Testing File Upload...")
    with open(FILE_PATH, "rb") as f:
        files = {"file": (os.path.basename(FILE_PATH), f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = requests.post(f"{API_URL}/upload", files=files)
        
    assert response.status_code == 200, f"Upload failed: {response.text}"
    meta = response.json()
    print("Upload Success!")
    print(f"Returned temporary filename: {meta['filename']}")
    print(f"Detected sheets: {meta['sheets']}")
    print(f"Detected columns: {meta['columns']}")
    
    filename = meta['filename']
    assert "Customers" in meta['sheets']
    assert "Name" in meta['columns']
    
    # 2. Test /search (case-insensitive, partial matching)
    print("\n2. Testing Search...")
    search_payload = {
        "filename": filename,
        "search_term": "Alice"
    }
    response = requests.post(f"{API_URL}/search", json=search_payload)
    assert response.status_code == 200, f"Search failed: {response.text}"
    search_results = response.json()
    print("Search Success!")
    print(f"Total Matches: {search_results['total_matches']}")
    print(f"Sheets Matched: {search_results['sheets_matched']}")
    print(f"Preview rows count: {len(search_results['preview_rows'])}")
    
    assert search_results['total_matches'] > 0
    assert "Customers" in search_results['sheets_matched']
    assert "Orders" in search_results['sheets_matched']
    
    # 3. Test /download
    print("\n3. Testing Download...")
    download_url = f"{API_URL}/download/{filename}/Alice"
    response = requests.get(download_url)
    assert response.status_code == 200, f"Download failed: {response.text}"
    
    download_path = "downloaded_test_result.xlsx"
    with open(download_path, "wb") as f:
        f.write(response.content)
    print(f"Downloaded results saved to: {download_path}")
    
    # 4. Verify spreadsheet file styling & structure
    print("\n4. Verifying spreadsheet file styling...")
    wb = load_workbook(download_path)
    ws = wb.active
    
    # Check headers
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
    assert "Source Sheet" in headers
    
    # Check style of header (row 1, cell 1)
    header_cell = ws.cell(row=1, column=1)
    print(f"Header cell font bold: {header_cell.font.bold}")
    print(f"Header cell fill color: {header_cell.fill.start_color.rgb}")
    assert header_cell.font.bold is True
    assert header_cell.fill.start_color.rgb == "00DCE6F1" or header_cell.fill.start_color.rgb == "DCE6F1"
    
    # Check style of Source Sheet column data cells (column 1, row 2)
    source_sheet_cell = ws.cell(row=2, column=1)
    print(f"Source Sheet data cell fill color: {source_sheet_cell.fill.start_color.rgb}")
    assert source_sheet_cell.fill.start_color.rgb == "00FFF2CC" or source_sheet_cell.fill.start_color.rgb == "FFF2CC"
    
    # Cleanup downloaded test file
    if os.path.exists(download_path):
        os.remove(download_path)
        
    print("\nAll integration tests passed successfully!")

if __name__ == "__main__":
    run_tests()
